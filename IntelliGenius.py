import time
import os
import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import PatternFill, Font

# Global variables
prop = {}
driver = None
book = None

# Load user configuration data
def load_user_data():
    global prop
    try:
        with open("Config.properties", "r") as file:
            for line in file:
                if '=' in line:
                    key, value = line.strip().split('=')
                    prop[key] = value
        print(f"Driver: {driver}")
    except Exception as e:
        print(f"Error loading user data: {e}")

# Initialize the WebDriver and open the URL
def launch_url():
    global driver
    chrome_driver_path = "C:\\Users\\ankits.jain_infobean\\Downloads\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe"
    options = Options()
    driver = webdriver.Chrome(service=Service(chrome_driver_path), options=options)
    driver.maximize_window()
    driver.delete_all_cookies()
    driver.get(prop.get("url"))
    driver.implicitly_wait(20)
    print("URL launched")
    print("Please login using username and password")

# Login to the dashboard
def login_to_dashboard():
    driver.find_element(By.ID, "username").send_keys(prop.get("username"))
    driver.find_element(By.ID, "password").send_keys(prop.get("password"))
    driver.find_element(By.XPATH, "//button[@class='login-submit-btn']").click()

    try:
        WebDriverWait(driver, 20).until(
            EC.visibility_of_all_elements(driver.find_elements(By.CSS_SELECTOR, ".navbar-brand"))
        )
    except Exception as e:
        print(f"Error during login: {e}")

    print("Navigated to Dashboard")

# Attach a file to the question prompt
def attach_question_file():
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, ".attach-file-icon svg").click()
    driver.find_element(By.XPATH, "//div[@class='select-doc-option ']").click()
    attach = driver.find_element(By.XPATH, "//div[@id='prompt-box']//form//input")
    attach.send_keys(prop.get("attachFile"))
    time.sleep(1)
    attach.send_keys(Keys.ARROW_DOWN, Keys.ENTER)
    time.sleep(1)
    driver.find_element(By.XPATH, "//div[@id='prompt-box']//form//button[text()='Select']").click()

# Wait until a specified element is invisible
def explicit_wait_invisible():
    try:
        WebDriverWait(driver, 300).until(EC.invisibility_of_element(driver.find_element(By.ID, "loader")))
    except Exception as e:
        print(f"Error in explicit wait: {e}")

# Enter a query into the input field
def enter_query(query):
    query_field = driver.find_element(By.XPATH, "//input[@class='form-control prompt']")
    query_field.clear()
    query_field.send_keys(query)
    query_field.send_keys(Keys.ENTER)
    time.sleep(2)
    explicit_wait_invisible()

# Save and retrieve results from the query
def save_result():
    result = driver.find_element(By.XPATH, "//div[@id='chatHistory']/div[last()]//div[@class='answer']").text
    print(f"Search result: {result}")
    print("="*80)
    return result

# Open the Excel file and sheet for reading
def open_reader(sheet_name):
    global book
    try:
        wb = openpyxl.load_workbook(file_path())
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"Error opening Excel file: {e}")

# Read the queries from the Excel sheet and perform the automated steps
def read_queries():
    sheet = open_reader(prop.get("sheetName"))
    row_count = sheet.max_row

    for i in range(1, row_count + 1):
        row = sheet[i]
        if i == 1:
            # Add a new header with the current date
            header = sheet.cell(row=1, column=2)
            header.value = f"Automated execution - Dated ({get_today_date()})"
            header.font = Font(bold=True)
            header.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            sheet.column_dimensions['B'].width = 30
        else:
            query = row[0].value
            print(f"***Entered query**** : {query}")
            enter_query(query)

            result = save_result()

            # Write the result in the next column
            cell = sheet.cell(row=i, column=2)
            cell.value = result
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Save the file after each iteration
        wb.save(file_path())

# Get the current date and time as a string
def get_today_date():
    return datetime.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")

# Get the file path for the Excel file
def file_path():
    return os.path.join(os.getcwd(), prop.get("fileName"))

# Main function to orchestrate the flow
def main():
    load_user_data()
    launch_url()
    login_to_dashboard()
    attach_question_file()
    read_queries()
    driver.quit()

# Run the program
if __name__ == "__main__":
    main()
